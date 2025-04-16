try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: Modul 'openpyxl' belum terinstall.")
    print("Silakan install terlebih dahulu dengan perintah:")
    print("pip install openpyxl")
    exit()

import os
from datetime import datetime

# Excel file paths
ZAKAT_DATA_FILE = "zakat_data.xlsx"
MASTER_BERAS_FILE = "master_beras.xlsx"
TRANSAKSI_ZAKAT_FILE = "transaksi_zakat.xlsx"

def initialize_files():
    """Initialize Excel files with headers if they don't exist"""
    try:
        if not os.path.exists(ZAKAT_DATA_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Zakat Data"
            ws.append(["ID", "Nama", "Jenis Zakat", "Jumlah", "Tanggal"])
            wb.save(ZAKAT_DATA_FILE)
        
        if not os.path.exists(MASTER_BERAS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Master Beras"
            ws.append(["ID", "Nama Beras", "Harga per Kg"])
            wb.save(MASTER_BERAS_FILE)
        
        if not os.path.exists(TRANSAKSI_ZAKAT_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Transaksi Zakat"
            ws.append(["ID", "ID Zakat", "ID Beras", "Jumlah Beras", "Total Harga", "Tanggal"])
            wb.save(TRANSAKSI_ZAKAT_FILE)
    except PermissionError:
        print("Error: Tidak bisa membuat file. Pastikan tidak ada file Excel yang sedang terbuka.")
        exit()
    except Exception as e:
        print(f"Error inisialisasi file: {str(e)}")
        exit()

def get_next_id(file_path):
    """Get the next available ID for a given Excel file"""
    try:
        if not os.path.exists(file_path):
            return 1
        
        wb = load_workbook(file_path)
        ws = wb.active
        max_id = 0
        
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row and row[0] is not None and isinstance(row[0], (int, float)):
                current_id = int(row[0])
                if current_id > max_id:
                    max_id = current_id
        
        return max_id + 1
    except Exception as e:
        print(f"Error mendapatkan ID: {str(e)}")
        return 1  # Return default ID if error occurs

def validate_date(date_str):
    """Validate date format (YYYY-MM-DD)"""
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def add_zakat(nama, jenis_zakat, jumlah, tanggal):
    """Add new zakat data to the Excel file"""
    try:
        # Input validation
        if not nama or not isinstance(nama, str):
            print("Error: Nama harus berupa teks dan tidak boleh kosong")
            return False
            
        if not jenis_zakat or not isinstance(jenis_zakat, str):
            print("Error: Jenis zakat harus berupa teks dan tidak boleh kosong")
            return False
            
        try:
            jumlah = float(jumlah)
            if jumlah <= 0:
                print("Error: Jumlah harus lebih besar dari 0")
                return False
        except (ValueError, TypeError):
            print("Error: Jumlah harus berupa angka")
            return False
            
        if not validate_date(tanggal):
            print("Error: Format tanggal tidak valid. Gunakan format YYYY-MM-DD")
            return False

        wb = load_workbook(ZAKAT_DATA_FILE)
        ws = wb.active
        new_id = get_next_id(ZAKAT_DATA_FILE)
        
        ws.append([new_id, nama.strip(), jenis_zakat.strip(), jumlah, tanggal])
        wb.save(ZAKAT_DATA_FILE)
        print(f"Data zakat berhasil ditambahkan dengan ID: {new_id}")
        return True
    except PermissionError:
        print("Error: File sedang digunakan. Tutup file Excel terlebih dahulu.")
        return False
    except Exception as e:
        print(f"Error menambahkan zakat: {str(e)}")
        return False

def update_zakat(id, nama, jenis_zakat, jumlah, tanggal):
    """Update existing zakat data"""
    try:
        # Input validation
        try:
            id = int(id)
            if id <= 0:
                print("Error: ID harus angka positif")
                return False
        except (ValueError, TypeError):
            print("Error: ID harus berupa angka")
            return False
            
        if not nama or not isinstance(nama, str):
            print("Error: Nama harus berupa teks dan tidak boleh kosong")
            return False
            
        if not jenis_zakat or not isinstance(jenis_zakat, str):
            print("Error: Jenis zakat harus berupa teks dan tidak boleh kosong")
            return False
            
        try:
            jumlah = float(jumlah)
            if jumlah <= 0:
                print("Error: Jumlah harus lebih besar dari 0")
                return False
        except (ValueError, TypeError):
            print("Error: Jumlah harus berupa angka")
            return False
            
        if not validate_date(tanggal):
            print("Error: Format tanggal tidak valid. Gunakan format YYYY-MM-DD")
            return False

        wb = load_workbook(ZAKAT_DATA_FILE)
        ws = wb.active
        found = False
        
        for row in ws.iter_rows(min_row=2):
            if row[0].value == id:
                row[1].value = nama.strip()
                row[2].value = jenis_zakat.strip()
                row[3].value = jumlah
                row[4].value = tanggal
                found = True
                break
        
        if found:
            wb.save(ZAKAT_DATA_FILE)
            print(f"Data zakat dengan ID {id} berhasil diperbarui")
            return True
        else:
            print(f"Error: ID {id} tidak ditemukan")
            return False
    except PermissionError:
        print("Error: File sedang digunakan. Tutup file Excel terlebih dahulu.")
        return False
    except Exception as e:
        print(f"Error memperbarui zakat: {str(e)}")
        return False

def delete_zakat(id):
    """Delete zakat data by ID"""
    try:
        # Input validation
        try:
            id = int(id)
            if id <= 0:
                print("Error: ID harus angka positif")
                return False
        except (ValueError, TypeError):
            print("Error: ID harus berupa angka")
            return False

        wb = load_workbook(ZAKAT_DATA_FILE)
        ws = wb.active
        rows_to_delete = []
        
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == id:
                rows_to_delete.append(idx)
        
        for idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(idx)
        
        if rows_to_delete:
            wb.save(ZAKAT_DATA_FILE)
            print(f"Data zakat dengan ID {id} berhasil dihapus")
            return True
        else:
            print(f"Error: ID {id} tidak ditemukan")
            return False
    except PermissionError:
        print("Error: File sedang digunakan. Tutup file Excel terlebih dahulu.")
        return False
    except Exception as e:
        print(f"Error menghapus zakat: {str(e)}")
        return False

def add_beras(nama_beras, harga_per_kg):
    """Add new beras data to the Excel file"""
    try:
        # Input validation
        if not nama_beras or not isinstance(nama_beras, str):
            print("Error: Nama beras harus berupa teks dan tidak boleh kosong")
            return False
            
        try:
            harga_per_kg = float(harga_per_kg)
            if harga_per_kg <= 0:
                print("Error: Harga harus lebih besar dari 0")
                return False
        except (ValueError, TypeError):
            print("Error: Harga harus berupa angka")
            return False

        wb = load_workbook(MASTER_BERAS_FILE)
        ws = wb.active
        new_id = get_next_id(MASTER_BERAS_FILE)
        
        ws.append([new_id, nama_beras.strip(), harga_per_kg])
        wb.save(MASTER_BERAS_FILE)
        print(f"Data beras berhasil ditambahkan dengan ID: {new_id}")
        return True
    except PermissionError:
        print("Error: File sedang digunakan. Tutup file Excel terlebih dahulu.")
        return False
    except Exception as e:
        print(f"Error menambahkan beras: {str(e)}")
        return False

def view_master_beras():
    """View all master beras data"""
    try:
        if not os.path.exists(MASTER_BERAS_FILE):
            print("Belum ada data master beras")
            return
        
        wb = load_workbook(MASTER_BERAS_FILE)
        ws = wb.active
        
        if ws.max_row <= 1:
            print("Belum ada data master beras")
            return
        
        print("\nMaster Data Beras:")
        print("-" * 50)
        print(f"{'ID':<5} | {'Nama Beras':<20} | {'Harga per Kg':<15}")
        print("-" * 50)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                print(f"{row[0]:<5} | {row[1]:<20} | {row[2]:<15.2f}")
    except Exception as e:
        print(f"Error menampilkan master beras: {str(e)}")

def add_transaksi_zakat(id_zakat, id_beras, jumlah_beras, tanggal):
    """Add new zakat transaction"""
    try:
        # Input validation
        try:
            id_zakat = int(id_zakat)
            if id_zakat <= 0:
                print("Error: ID zakat harus angka positif")
                return False
        except (ValueError, TypeError):
            print("Error: ID zakat harus berupa angka")
            return False
            
        try:
            id_beras = int(id_beras)
            if id_beras <= 0:
                print("Error: ID beras harus angka positif")
                return False
        except (ValueError, TypeError):
            print("Error: ID beras harus berupa angka")
            return False
            
        try:
            jumlah_beras = float(jumlah_beras)
            if jumlah_beras <= 0:
                print("Error: Jumlah beras harus lebih besar dari 0")
                return False
        except (ValueError, TypeError):
            print("Error: Jumlah beras harus berupa angka")
            return False
            
        if not validate_date(tanggal):
            print("Error: Format tanggal tidak valid. Gunakan format YYYY-MM-DD")
            return False

        # Check if zakat ID exists
        zakat_exists = False
        zakat_name = ""
        if os.path.exists(ZAKAT_DATA_FILE):
            wb_zakat = load_workbook(ZAKAT_DATA_FILE)
            ws_zakat = wb_zakat.active
            for row in ws_zakat.iter_rows(min_row=2, values_only=True):
                if row and row[0] == id_zakat:
                    zakat_exists = True
                    zakat_name = row[1]
                    break
        
        if not zakat_exists:
            print(f"Error: ID zakat {id_zakat} tidak ditemukan!")
            return False
        
        # Check if beras ID exists and get price
        beras_price = None
        beras_name = ""
        if os.path.exists(MASTER_BERAS_FILE):
            wb_beras = load_workbook(MASTER_BERAS_FILE)
            ws_beras = wb_beras.active
            for row in ws_beras.iter_rows(min_row=2, values_only=True):
                if row and row[0] == id_beras:
                    beras_price = row[2]
                    beras_name = row[1]
                    break
        
        if beras_price is None:
            print(f"Error: ID beras {id_beras} tidak ditemukan!")
            return False
        
        total_harga = beras_price * jumlah_beras
        
        # Add transaction
        wb = load_workbook(TRANSAKSI_ZAKAT_FILE)
        ws = wb.active
        new_id = get_next_id(TRANSAKSI_ZAKAT_FILE)
        
        ws.append([new_id, id_zakat, id_beras, jumlah_beras, total_harga, tanggal])
        wb.save(TRANSAKSI_ZAKAT_FILE)
        
        print("\nTransaksi zakat berhasil ditambahkan:")
        print(f"ID Transaksi: {new_id}")
        print(f"Nama Zakat: {zakat_name}")
        print(f"Jenis Beras: {beras_name}")
        print(f"Jumlah Beras: {jumlah_beras} kg")
        print(f"Total Harga: Rp {total_harga:,.2f}")
        print(f"Tanggal: {tanggal}")
        
        return True
    except PermissionError:
        print("Error: File sedang digunakan. Tutup file Excel terlebih dahulu.")
        return False
    except Exception as e:
        print(f"Error menambahkan transaksi: {str(e)}")
        return False

def view_transaksi_zakat():
    """View all zakat transactions"""
    try:
        if not os.path.exists(TRANSAKSI_ZAKAT_FILE):
            print("Belum ada data transaksi zakat")
            return
        
        wb_trans = load_workbook(TRANSAKSI_ZAKAT_FILE)
        ws_trans = wb_trans.active
        
        if ws_trans.max_row <= 1:
            print("Belum ada data transaksi zakat")
            return
        
        # Load zakat data
        zakat_data = {}
        if os.path.exists(ZAKAT_DATA_FILE):
            wb_zakat = load_workbook(ZAKAT_DATA_FILE)
            ws_zakat = wb_zakat.active
            for row in ws_zakat.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    zakat_data[row[0]] = (row[1], row[2])  # (nama, jenis_zakat)
        
        # Load beras data
        beras_data = {}
        if os.path.exists(MASTER_BERAS_FILE):
            wb_beras = load_workbook(MASTER_BERAS_FILE)
            ws_beras = wb_beras.active
            for row in ws_beras.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    beras_data[row[0]] = row[1]  # nama_beras
        
        print("\nDaftar Transaksi Zakat:")
        print("-" * 120)
        print(f"{'ID':<5} | {'Nama':<20} | {'Jenis Zakat':<15} | {'Beras':<15} | {'Jumlah (kg)':<10} | {'Total Harga':<15} | {'Tanggal':<10}")
        print("-" * 120)
        
        for row in ws_trans.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                zakat_info = zakat_data.get(row[1], ("Unknown", "Unknown"))
                beras_name = beras_data.get(row[2], "Unknown")
                
                print(f"{row[0]:<5} | {zakat_info[0]:<20} | {zakat_info[1]:<15} | {beras_name:<15} | "
                      f"{row[3]:<10.2f} | Rp {row[4]:<12.2f} | {row[5]:<10}")
    except Exception as e:
        print(f"Error menampilkan transaksi: {str(e)}")

def export_to_excel():
    """Export zakat data to a new Excel file"""
    try:
        if not os.path.exists(ZAKAT_DATA_FILE):
            print("Tidak ada data zakat untuk diekspor")
            return
        
        wb = load_workbook(ZAKAT_DATA_FILE)
        ws = wb.active
        
        if ws.max_row <= 1:
            print("Tidak ada data zakat untuk diekspor")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"data_zakat_export_{timestamp}.xlsx"
        
        export_wb = Workbook()
        export_ws = export_wb.active
        export_ws.title = "Zakat Data Export"
        
        # Copy headers
        for row in ws.iter_rows(max_row=1):
            export_ws.append([cell.value for cell in row])
        
        # Copy data
        for row in ws.iter_rows(min_row=2):
            export_ws.append([cell.value for cell in row])
        
        export_wb.save(filename)
        print(f"\nData zakat berhasil diekspor ke file: {filename}")
        print(f"Lokasi file: {os.path.abspath(filename)}")
    except PermissionError:
        print("Error: Tidak bisa menulis file. Pastikan tidak ada file dengan nama yang sama yang sedang terbuka.")
    except Exception as e:
        print(f"Error ekspor data: {str(e)}")

def input_master_beras():
    """Input new master beras data from user"""
    print("\nTambah Data Master Beras")
    while True:
        nama_beras = input("Masukkan nama jenis beras (min 3 karakter): ").strip()
        if len(nama_beras) >= 3:
            break
        print("Error: Nama beras harus minimal 3 karakter")
    
    while True:
        harga_input = input("Masukkan harga per kg (contoh: 15000): ").strip()
        try:
            harga_per_kg = float(harga_input)
            if harga_per_kg > 0:
                break
            print("Error: Harga harus lebih besar dari 0")
        except ValueError:
            print("Error: Masukkan angka yang valid")
    
    if add_beras(nama_beras, harga_per_kg):
        print("Data master beras berhasil ditambahkan!")

def input_zakat_data():
    """Input zakat data from user"""
    print("\nTambah Data Zakat")
    while True:
        nama = input("Masukkan nama (min 3 karakter): ").strip()
        if len(nama) >= 3:
            break
        print("Error: Nama harus minimal 3 karakter")
    
    while True:
        jenis_zakat = input("Masukkan jenis zakat (contoh: Fitrah, Maal): ").strip()
        if jenis_zakat:
            break
        print("Error: Jenis zakat tidak boleh kosong")
    
    while True:
        jumlah_input = input("Masukkan jumlah zakat: ").strip()
        try:
            jumlah = float(jumlah_input)
            if jumlah > 0:
                break
            print("Error: Jumlah harus lebih besar dari 0")
        except ValueError:
            print("Error: Masukkan angka yang valid")
    
    while True:
        tanggal = input("Masukkan tanggal (YYYY-MM-DD): ").strip()
        if validate_date(tanggal):
            break
        print("Error: Format tanggal tidak valid. Gunakan format YYYY-MM-DD")
    
    if add_zakat(nama, jenis_zakat, jumlah, tanggal):
        print("Data zakat berhasil ditambahkan!")

def input_transaksi_zakat():
    """Input zakat transaction from user"""
    print("\nTambah Transaksi Zakat")
    
    # Show available zakat data
    try:
        if os.path.exists(ZAKAT_DATA_FILE):
            wb = load_workbook(ZAKAT_DATA_FILE)
            ws = wb.active
            if ws.max_row > 1:
                print("\nDaftar Zakat Tersedia:")
                print("-" * 60)
                print(f"{'ID':<5} | {'Nama':<20} | {'Jenis Zakat':<15} | {'Jumlah':<10}")
                print("-" * 60)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] is not None:
                        print(f"{row[0]:<5} | {row[1]:<20} | {row[2]:<15} | {row[3]:<10}")
    except Exception:
        pass
    
    while True:
        id_zakat_input = input("\nMasukkan ID zakat: ").strip()
        try:
            id_zakat = int(id_zakat_input)
            if id_zakat > 0:
                break
            print("Error: ID harus lebih besar dari 0")
        except ValueError:
            print("Error: Masukkan angka ID yang valid")
    
    # Show available beras data
    view_master_beras()
    
    while True:
        id_beras_input = input("\nMasukkan ID beras: ").strip()
        try:
            id_beras = int(id_beras_input)
            if id_beras > 0:
                break
            print("Error: ID harus lebih besar dari 0")
        except ValueError:
            print("Error: Masukkan angka ID yang valid")
    
    while True:
        jumlah_beras_input = input("Masukkan jumlah beras (kg): ").strip()
        try:
            jumlah_beras = float(jumlah_beras_input)
            if jumlah_beras > 0:
                break
            print("Error: Jumlah harus lebih besar dari 0")
        except ValueError:
            print("Error: Masukkan angka yang valid")
    
    while True:
        tanggal = input("Masukkan tanggal transaksi (YYYY-MM-DD): ").strip()
        if validate_date(tanggal):
            break
        print("Error: Format tanggal tidak valid. Gunakan format YYYY-MM-DD")
    
    add_transaksi_zakat(id_zakat, id_beras, jumlah_beras, tanggal)

def main_menu():
    """Display main menu"""
    print("\n" + "="*50)
    print("SISTEM MANAJEMEN ZAKAT".center(50))
    print("="*50)
    print("1. Kelola Data Zakat")
    print("2. Kelola Master Beras")
    print("3. Kelola Transaksi Zakat")
    print("4. Ekspor Data")
    print("5. Keluar")

def zakat_menu():
    """Display zakat management menu"""
    print("\n" + "="*50)
    print("KELOLA DATA ZAKAT".center(50))
    print("="*50)
    print("1. Tambah Data Zakat")
    print("2. Edit Data Zakat")
    print("3. Hapus Data Zakat")
    print("4. Kembali ke Menu Utama")

def beras_menu():
    """Display beras management menu"""
    print("\n" + "="*50)
    print("KELOLA MASTER BERAS".center(50))
    print("="*50)
    print("1. Tambah Data Beras")
    print("2. Lihat Data Beras")
    print("3. Kembali ke Menu Utama")

def transaksi_menu():
    """Display transaction management menu"""
    print("\n" + "="*50)
    print("KELOLA TRANSAKSI ZAKAT".center(50))
    print("="*50)
    print("1. Tambah Transaksi Zakat")
    print("2. Lihat Transaksi Zakat")
    print("3. Kembali ke Menu Utama")

def main():
    initialize_files()
    
    while True:
        main_menu()
        choice = input("Pilih menu (1-5): ").strip()
        
        if choice == "1":  # Kelola Data Zakat
            while True:
                zakat_menu()
                sub_choice = input("Pilih opsi (1-4): ").strip()
                
                if sub_choice == "1":  # Tambah Data Zakat
                    input_zakat_data()
                elif sub_choice == "2":  # Edit Data Zakat
                    try:
                        id_zakat = input("Masukkan ID zakat yang akan diedit: ").strip()
                        nama = input("Masukkan nama baru: ").strip()
                        jenis_zakat = input("Masukkan jenis zakat baru: ").strip()
                        jumlah = input("Masukkan jumlah baru: ").strip()
                        tanggal = input("Masukkan tanggal baru (YYYY-MM-DD): ").strip()
                        update_zakat(id_zakat, nama, jenis_zakat, jumlah, tanggal)
                    except Exception as e:
                        print(f"Error: {str(e)}")
                elif sub_choice == "3":  # Hapus Data Zakat
                    try:
                        id_zakat = input("Masukkan ID zakat yang akan dihapus: ").strip()
                        delete_zakat(id_zakat)
                    except Exception as e:
                        print(f"Error: {str(e)}")
                elif sub_choice == "4":  # Kembali
                    break
                else:
                    print("Pilihan tidak valid. Silakan coba lagi.")
        
        elif choice == "2":  # Kelola Master Beras
            while True:
                beras_menu()
                sub_choice = input("Pilih opsi (1-3): ").strip()
                
                if sub_choice == "1":  # Tambah Data Beras
                    input_master_beras()
                elif sub_choice == "2":  # Lihat Data Beras
                    view_master_beras()
                elif sub_choice == "3":  # Kembali
                    break
                else:
                    print("Pilihan tidak valid. Silakan coba lagi.")
        
        elif choice == "3":  # Kelola Transaksi Zakat
            while True:
                transaksi_menu()
                sub_choice = input("Pilih opsi (1-3): ").strip()
                
                if sub_choice == "1":  # Tambah Transaksi
                    input_transaksi_zakat()
                elif sub_choice == "2":  # Lihat Transaksi
                    view_transaksi_zakat()
                elif sub_choice == "3":  # Kembali
                    break
                else:
                    print("Pilihan tidak valid. Silakan coba lagi.")
        
        elif choice == "4":  # Ekspor Data
            export_to_excel()
        
        elif choice == "5":  # Keluar
            print("Terima kasih telah menggunakan Sistem Manajemen Zakat.")
            break
        
        else:
            print("Pilihan tidak valid. Silakan pilih 1-5.")

if __name__ == "__main__":
    main()